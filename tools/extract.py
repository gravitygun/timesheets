from openpyxl import load_workbook
import json

def extract_excel_structure(filepath):
    # Load with data_only=False to get formulas
    wb_formulas = load_workbook(filepath, data_only=False)
    # Load again with data_only=True to get computed values
    wb_values = load_workbook(filepath, data_only=True)
    
    result = {}
    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]
        
        sheet_data = {}
        for row in ws_f.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_ref = cell.coordinate
                    sheet_data[cell_ref] = {
                        "formula": str(cell.value) if str(cell.value).startswith("=") else None,
                        "value": ws_v[cell_ref].value,
                        "type": type(ws_v[cell_ref].value).__name__
                    }
        result[sheet_name] = sheet_data
    
    return result

# Export to JSON for review
data = extract_excel_structure("/Users/wayne/OneDrive/Blitterbyte/Timesheet 2025.xlsx")
with open("spreadsheet_structure.json", "w") as f:
    json.dump(data, f, indent=2, default=str)
